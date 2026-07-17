#Import DB credentials
from dotenv import load_dotenv
import os

load_dotenv()



# pip install mysql-connector-python, pip install openpyxl
#import load_workbook from openpyxl to load the Excel file
from openpyxl import load_workbook

#import mysql.connector to connect with mysql 
import mysql.connector

# Connect with sql 
connection = mysql.connector.connect(
    host='localhost', 
    user='root', 
    password='Anupam@2002', 
    database='secure'
);

cursor = connection.cursor();
print(cursor)

sql = "INSERT IGNORE INTO batch_number (sl_no, batch_no, mc_no, created_on, shift, design_name, product_colour, logo, raw_material_code, raw_material_supplier, master_batch_code, master_batch_supplier, master_batch_color, master_batch_percentage, mc_speed, created_by, remarks) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

#Loading data from our excel file 
sheet = load_workbook(filename='BATCH_NO_FILE.xlsx').active
#Created function that iterate on row and get all value 
for row in sheet.iter_rows(min_row=10, max_col=17, values_only=True):

    (
        sl_no,
        batch_no,
        mc_no,
        created_on,
        shift,
        design_name,
        product_colour,
        logo,
        raw_material_code,
        raw_material_supplier,
        master_batch_code,
        master_batch_supplier,
        master_batch_color,
        master_batch_percentage,
        mc_speed,
        created_by,
        remarks
    ) = row

    # Skip rows with empty batch number
    if batch_no is None:
        continue
    
    
    values = (
        sl_no,
        batch_no,
        mc_no,
        created_on,
        shift,
        design_name,
        product_colour,
        logo,
        raw_material_code,
        raw_material_supplier,
        master_batch_code,
        master_batch_supplier,
        master_batch_color,
        master_batch_percentage,
        mc_speed,
        created_by,
        remarks
    )
    cursor.execute(sql, values)

    

# Commit once after all inserts
connection.commit()

# import tkinter as tk
# from tkinter import messagebox

# root = tk.Tk()
# root.withdraw()  # Hide the main window

# messagebox.showinfo("Python", "Hello World!\nPython script executed successfully.")

